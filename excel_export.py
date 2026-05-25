"""
Excel Export Module
Creates a professionally formatted workbook with:
- Sheet 1: Deal Dashboard (sorted by opportunity)
- Sheet 2: Reference Prices (full database)
- Sheet 3: Scrape Log (metadata)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from datetime import datetime, timezone


# Color scheme
DARK_BG = "1B2631"
HEADER_BG = "2C3E50"
ACCENT_GREEN = "27AE60"
ACCENT_RED = "E74C3C"
ACCENT_ORANGE = "F39C12"
ACCENT_BLUE = "2980B9"
LIGHT_ROW = "F8F9FA"
WHITE = "FFFFFF"
DARK_TEXT = "1B2631"

THIN_BORDER = Border(
    left=Side(style="thin", color="D5D8DC"),
    right=Side(style="thin", color="D5D8DC"),
    top=Side(style="thin", color="D5D8DC"),
    bottom=Side(style="thin", color="D5D8DC"),
)


def export_to_excel(listings: list[dict], filepath: str):
    """Export enriched listings to a formatted Excel workbook."""
    wb = openpyxl.Workbook()

    _create_dashboard_sheet(wb, listings)
    _create_reference_sheet(wb)
    _create_log_sheet(wb, listings)

    wb.save(filepath)


def _create_dashboard_sheet(wb: openpyxl.Workbook, listings: list[dict]):
    """Main deals dashboard."""
    ws = wb.active
    ws.title = "☕ Deals Dashboard"

    # Column definitions
    columns = [
        ("Rating", 16),
        ("Brand", 18),
        ("Model", 22),
        ("Source", 16),
        ("Current Bid", 14),
        ("Retail Price", 14),
        ("Discount %", 12),
        ("Gross Potential", 16),
        ("Est. Resale", 14),
        ("Net Profit Est.", 16),
        ("Category", 16),
        ("Lot #", 10),
        ("End Time", 20),
        ("Location", 20),
        ("Condition", 22),
        ("URL", 50),
        ("Listing Title", 60),
        ("Scraped At", 22),
    ]

    # Set column widths
    for i, (name, width) in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Title row
    ws.merge_cells("A1:R1")
    title_cell = ws["A1"]
    title_cell.value = f"☕ Coffee Machine Arbitrage Tracker — {datetime.now().strftime('%d %b %Y %H:%M')}"
    title_cell.font = Font(name="Calibri", size=16, bold=True, color=WHITE)
    title_cell.fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 40

    # Subtitle
    ws.merge_cells("A2:R2")
    sub_cell = ws["A2"]
    sub_cell.value = f"Sources: John Pye · i-bidder · Bidspotter · BPI Auctions · Pro Auction · Wilsons · Auction News  |  {len(listings)} listings found"
    sub_cell.font = Font(name="Calibri", size=10, color="95A5A6")
    sub_cell.fill = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")
    ws.row_dimensions[2].height = 24

    # Header row
    header_font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, (name, _) in enumerate(columns, 1):
        cell = ws.cell(row=3, column=i, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = THIN_BORDER
    ws.row_dimensions[3].height = 30

    # Data rows
    for row_idx, item in enumerate(listings, 4):
        # Alternate row coloring
        row_fill = PatternFill(start_color=LIGHT_ROW, end_color=LIGHT_ROW, fill_type="solid") if row_idx % 2 == 0 else None

        values = [
            item.get("rating", ""),
            item.get("matched_brand", ""),
            item.get("matched_model", ""),
            item.get("source", ""),
            item.get("current_bid"),
            item.get("retail_price_ref"),
            item.get("discount_pct"),
            item.get("gross_potential"),
            item.get("conservative_resale"),
            item.get("conservative_profit"),
            item.get("category", ""),
            item.get("lot_number", ""),
            item.get("end_time", ""),
            item.get("location", ""),
            item.get("condition", ""),
            item.get("url", ""),
            item.get("title", ""),
            item.get("scraped_at", ""),
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Calibri", size=10, color=DARK_TEXT)
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border = THIN_BORDER

            if row_fill:
                cell.fill = row_fill

            # Format currency columns
            if col_idx in (5, 6, 8, 9, 10) and isinstance(value, (int, float)):
                cell.number_format = '£#,##0.00'

            # Format percentage
            if col_idx == 7 and isinstance(value, (int, float)):
                cell.number_format = '0.0"%"'

            # Color-code the rating column
            if col_idx == 1:
                if "STRONG BUY" in str(value):
                    cell.font = Font(name="Calibri", size=10, bold=True, color=ACCENT_GREEN)
                elif "GOOD DEAL" in str(value):
                    cell.font = Font(name="Calibri", size=10, bold=True, color=ACCENT_BLUE)
                elif "FAIR" in str(value):
                    cell.font = Font(name="Calibri", size=10, color=ACCENT_ORANGE)
                elif "OVERPRICED" in str(value):
                    cell.font = Font(name="Calibri", size=10, color=ACCENT_RED)

            # Make URL a hyperlink
            if col_idx == 16 and value and str(value).startswith("http"):
                cell.hyperlink = str(value)
                cell.font = Font(name="Calibri", size=10, color=ACCENT_BLUE, underline="single")

    # Freeze panes (header row visible when scrolling)
    ws.freeze_panes = "A4"

    # Auto-filter
    ws.auto_filter.ref = f"A3:R{3 + len(listings)}"


def _create_reference_sheet(wb: openpyxl.Workbook):
    """Reference price database sheet."""
    from reference_prices import REFERENCE_MACHINES

    ws = wb.create_sheet("📊 Reference Prices")

    # Header
    headers = ["Search Key", "Brand", "Model", "Retail Price (£)", "Category"]
    header_font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 18

    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # Data
    for row_idx, (key, ref) in enumerate(sorted(REFERENCE_MACHINES.items()), 2):
        values = [key, ref["brand"], ref["model"], ref["retail_gbp"], ref["category"]]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Calibri", size=10)
            cell.border = THIN_BORDER
            if col_idx == 4:
                cell.number_format = '£#,##0'

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{1 + len(REFERENCE_MACHINES)}"


def _create_log_sheet(wb: openpyxl.Workbook, listings: list[dict]):
    """Scrape metadata and stats."""
    ws = wb.create_sheet("📋 Scrape Log")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50

    header_font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")

    for i, h in enumerate(["Metric", "Value"], 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = THIN_BORDER

    # Compute stats
    sources = {}
    for item in listings:
        src = item.get("source", "Unknown")
        sources[src] = sources.get(src, 0) + 1

    strong_buys = sum(1 for l in listings if l.get("rating") == "🔥 STRONG BUY")
    good_deals = sum(1 for l in listings if l.get("rating") == "✅ GOOD DEAL")
    manual_checks = sum(1 for l in listings if l.get("rating") == "🔍 MANUAL CHECK")

    stats = [
        ("Scrape Timestamp (UTC)", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")),
        ("Total Listings Found", len(listings)),
        ("🔥 Strong Buys (60%+ discount)", strong_buys),
        ("✅ Good Deals (40-60% discount)", good_deals),
        ("🔍 Manual Check Required", manual_checks),
        ("", ""),
        ("LISTINGS BY SOURCE", ""),
    ]
    for src, count in sorted(sources.items()):
        stats.append((f"  {src}", count))

    stats.extend([
        ("", ""),
        ("ABOUT THIS TOOL", ""),
        ("Strategy", "Monitor low-competition UK auction sites for premium coffee equipment"),
        ("Target Discount", "40%+ below retail = worth bidding"),
        ("Conservative Resale", "Assumes 65% of retail for refurbished resale"),
        ("Schedule", "Recommended: run at 09:00 and 21:00 daily"),
        ("Sources", "John Pye, i-bidder, Bidspotter, BPI, Pro Auction, Wilsons, Auction News"),
    ])

    for row_idx, (metric, value) in enumerate(stats, 2):
        cell_a = ws.cell(row=row_idx, column=1, value=metric)
        cell_b = ws.cell(row=row_idx, column=2, value=value)
        cell_a.font = Font(name="Calibri", size=10, bold=bool(metric and not metric.startswith(" ")))
        cell_b.font = Font(name="Calibri", size=10)
        cell_a.border = THIN_BORDER
        cell_b.border = THIN_BORDER
